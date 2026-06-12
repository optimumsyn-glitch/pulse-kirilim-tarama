import yfinance as yf
import pandas as pd
import numpy as np
import warnings
from datetime import datetime
import time
import os

warnings.filterwarnings("ignore")

# ==================== PİNESCRİPT AYARLARI ====================
left_len = right_len = 5
donchian_period = 20
rsi_length = 14
rsi_overbought = 70
rsi_oversold = 30
atr_period = 14
volume_lookback = 20
adx_length = 14
BREAKOUT_MULTIPLIER = 1.02
MIN_STRENGTH = 2.5

broken_res = {}

# ==================== PERİYOT SEÇİMİ ====================
print("=== Pulse Kırılım Tarayıcı ===")
print("Zaman Dilimi Seçin:")
print("1 → 1 Saat")
print("2 → 4 Saat")
print("3 → Günlük")
print("4 → Haftalık")
print("5 → Aylık")

secim = input("Seçiminiz (1-5): ").strip()
tf_map = {
    "1": {"name": "1h",  "interval": "1h",  "period": "60d"},
    "2": {"name": "4h",  "interval": "1h",  "period": "60d"},
    "3": {"name": "daily","interval": "1d",  "period": "120d"},
    "4": {"name": "weekly","interval": "1wk", "period": "3y"},
    "5": {"name": "monthly","interval": "1mo", "period": "5y"}
}
tf = tf_map.get(secim, {"name": "daily", "interval": "1d", "period": "120d"})
print(f"\nSeçilen Periyot: {tf['name'].upper()}\n")

# ==================== HİSSE LİSTESİ ====================
symbols = [
    'A1CAP', 'A1YEN', 'ACSEL', 'ADEL', 'ADESE', 'ADGYO', 'AEFES', 'AFYON', 'AGESA', 'AGHOL', 'AGROT', 'AGYO', 'AHGAZ', 'AHSGY', 'AKBNK', 'AKCNS', 'AKENR', 'AKFGY', 'AKFIS', 'AKFYE', 'AKGRT', 'AKMGY', 'AKSA', 'AKSEN', 'AKSUE', 'AKYHO', 'ALARK', 'ALCAR', 'ALCTL', 'ALFAS', 'ALGYO', 'ALKA', 'ALKIM', 'ALKLC', 'ALTNY', 'ALVES', 'ANELE', 'ANGEN', 'ANHYT', 'APBDL', 'APLIB', 'APMDL', 'APX30', 'ARASE', 'ARCLK', 'ARDYZ', 'ARENA', 'ARMGD', 'ARSAN', 'ARTMS', 'ARZUM', 'ASELS', 'ASGYO', 'ASTOR', 'ASUZU', 'ATAKP', 'ATATP', 'ATEKS', 'ATLAS', 'ATSYH', 'AVGYO', 'AVHOL', 'AVOD', 'AVPGY', 'AVTUR', 'AYCES', 'AYDEM', 'AYEN', 'AYES', 'AYGAZ', 'AZTEK', 'BAGFS', 'BAHKM', 'BAKAB', 'BALAT', 'BALSU', 'BANVT', 'BARMA', 'BASCM', 'BASGZ', 'BAYRK', 'BEGYO', 'BERA', 'BESLR', 'BEYAZ', 'BFREN', 'BIENY', 'BIGCH', 'BIGEN', 'BIMAS', 'BINBN', 'BINHO', 'BIOEN', 'BIZIM', 'BJKAS', 'BLCYT', 'BLUME', 'BMSCH', 'BMSTL', 'BNTAS', 'BOBET', 'BORLS', 'BORSK', 'BOSSA', 'BRISA', 'BRKSN', 'BRKVY', 'BRLSM', 'BRMEN', 'BRSAN', 'BRYAT', 'BSOKE', 'BTCIM', 'BUCIM', 'BULGS', 'BURCE', 'BURVA', 'BVSAN', 'BYDNR', 'CANTE', 'CASA', 'CATES', 'CCOLA', 'CELHA', 'CEMAS', 'CEMTS', 'CEMZY', 'CEOEM', 'CGCAM', 'CIMSA', 'CLEBI', 'CMBTN', 'CMENT', 'CONSE', 'COSMO', 'CRDFA', 'CRFSA', 'CUSAN', 'CVKMD', 'CWENE', 'DAGI', 'DAPGM', 'DARDL', 'DCTTR', 'DERHL', 'DERIM', 'DESA', 'DESPC', 'DEVA', 'DGATE', 'DGGYO', 'DGNMO', 'DIRIT', 'DITAS', 'DMRGD', 'DMSAS', 'DNISI', 'DOAS', 'DOBUR', 'DOCO', 'DOFER', 'DOFRB', 'DOGUB', 'DOHOL', 'DOKTA', 'DSTKF', 'DUNYH', 'DURDO', 'DURKN', 'DYOBY', 'DZGYO', 'EBEBK', 'ECILC', 'ECZYT', 'EDATA', 'EDIP', 'EFORC', 'EGEEN', 'EGEGY', 'EGEPO', 'EGGUB', 'EGPRO', 'EGSER', 'EKGYO', 'EKIZ', 'EKOS', 'EKSUN', 'ELITE', 'EMKEL', 'EMNIS', 'ENDAE', 'ENERY', 'ENJSA', 'ENKAI', 'ENSRI', 'ENTRA', 'EPLAS', 'ERBOS', 'ERCB', 'EREGL', 'ERSU', 'ESCAR', 'ESCOM', 'ESEN', 'ETILR', 'ETYAT', 'EUKYO', 'EUPWR', 'EUREN', 'EUYO', 'EYGYO', 'FENER', 'FLAP', 'FMIZP', 'FONET', 'FORTE', 'FRIGO', 'FZLGY', 'GARAN', 'GARFA', 'GEDIK', 'GEDZA', 'GENIL', 'GENTS', 'GEREL', 'GESAN', 'GLBMD', 'GLCVY', 'GLDTR', 'GLRMK', 'GLRYH', 'GMSTR', 'GMTAS', 'GOKNR', 'GOLTS', 'GOODY', 'GOZDE', 'GRNYO', 'GRSEL', 'GRTHO', 'GSDDE', 'GSDHO', 'GSRAY', 'GUBRF', 'GUNDG', 'GWIND', 'GZNMI', 'HALKB', 'HATEK', 'HATSN', 'HDFGS', 'HEDEF', 'HEKTS', 'HKTM', 'HLGYO', 'HOROZ', 'HRKET', 'HTTBT', 'HUBVC', 'HUNER', 'HURGZ', 'ICBCT', 'ICUGS', 'IDGYO', 'IEYHO', 'IHAAS', 'IHEVA', 'IHGZT', 'IHLAS', 'IHLGM', 'IHYAY', 'IMASM', 'INDES', 'INFO', 'INGRM', 'INTEK', 'INTEM', 'INVEO', 'INVES', 'IPEKE', 'ISBIR', 'ISBTR', 'ISCTR', 'ISDMR', 'ISFIN', 'ISGLK', 'ISGSY', 'ISGYO', 'ISKPL', 'ISMEN', 'ISSEN', 'IZENR', 'IZFAS', 'IZINV', 'IZMDC', 'JANTS', 'KAPLM', 'KAREL', 'KARSN', 'KARTN', 'KATMR', 'KAYSE', 'KBORU', 'KCAER', 'KCHOL', 'KENT', 'KERVN', 'KFEIN', 'KGYO', 'KIMMR', 'KLGYO', 'KLKIM', 'KLMSN', 'KLRHO', 'KLSER', 'KLSYN', 'KLYPV', 'KMPUR', 'KNFRT', 'KOCMT', 'KONKA', 'KONTR', 'KONYA', 'KOPOL', 'KORDS', 'KOTON', 'KOZAA', 'KOZAL', 'KRDMA', 'KRDMB', 'KRGYO', 'KRONT', 'KRPLS', 'KRSTL', 'KRTEK', 'KRVGD', 'KSTUR', 'KTLEV', 'KTSKR', 'KUTPO', 'KUVVA', 'KUYAS', 'KZBGY', 'KZGYO', 'LIDER', 'LILAK', 'LKMNH', 'LMKDC', 'LRSHO', 'LUKSK', 'LYDHO', 'LYDYE', 'MAALT', 'MACKO', 'MAGEN', 'MAKIM', 'MAKTK', 'MANAS', 'MARBL', 'MARKA', 'MARMR', 'MARTI', 'MAVI', 'MEDTR', 'MEGAP', 'MEGMT', 'MEKAG', 'MEPET', 'MERCN', 'MERIT', 'MERKO', 'METRO', 'MGROS', 'MHRGY', 'MIATK', 'MMCAS', 'MNDRS', 'MNDTR', 'MOBTL', 'MOGAN', 'MOPAS', 'MPARK', 'MRGYO', 'MRSHL', 'MSGYO', 'MTRKS', 'MTRYO', 'NATEN', 'NETAS', 'NIBAS', 'NTGAZ', 'NTHOL', 'NUGYO', 'NUHCM', 'OBAMS', 'OBASE', 'ODAS', 'ODINE', 'OFSYM', 'ONCSM', 'ONRYT', 'OPK30', 'OPT25', 'OPTGY', 'OPTLR', 'OPX30', 'ORCAY', 'ORGE', 'ORMA', 'OSMEN', 'OSTIM', 'OTKAR', 'OTTO', 'OYAKC', 'OYAYO', 'OYLUM', 'OYYAT', 'OZATD', 'OZGYO', 'OZKGY', 'OZRDN', 'OZSUB', 'OZYSR', 'PAGYO', 'PAMEL', 'PAPIL', 'PARSN', 'PASEU', 'PATEK', 'PCILT', 'PEKGY', 'PENGD', 'PENTA', 'PETKM', 'PETUN', 'PGSUS', 'PINSU', 'PKART', 'PKENT', 'PLTUR', 'PNLSN', 'PNSUT', 'POLHO', 'POLTK', 'PRDGS', 'PRKAB', 'PRKME', 'PSDTC', 'PSGYO', 'QNBFK', 'QNBTR', 'QTEMZ', 'QUAGR', 'RALYH', 'RAYSG', 'REEDR', 'RGYAS', 'RNPOL', 'RTALB', 'RUBNS', 'RUZYE', 'RYGYO', 'RYSAS', 'SAFKR', 'SAHOL', 'SAMAT', 'SANEL', 'SANFM', 'SANKO', 'SARKY', 'SASA', 'SDTTR', 'SEGMN', 'SEGYO', 'SEKFK', 'SEKUR', 'SELEC', 'SELVA', 'SERNT', 'SILVR', 'SISE', 'SKBNK', 'SKTAS', 'SKYLP', 'SKYMD', 'SMART', 'SMRTG', 'SMRVA', 'SNGYO', 'SNICA', 'SNKRN', 'SNPAM', 'SODSN', 'SOKE', 'SOKM', 'SONME', 'SRVGY', 'SUMAS', 'SUNTK', 'SURGY', 'SUWEN', 'TABGD', 'TARKM', 'TATEN', 'TATGD', 'TAVHL', 'TBORG', 'TCELL', 'TCKRC', 'TDGYO', 'TEHOL', 'TEKTU', 'TERA', 'TEZOL', 'TGSAS', 'THYAO', 'TKFEN', 'TKNSA', 'TLMAN', 'TMSN', 'TNZTP', 'TOASO', 'TRCAS', 'TRGYO', 'TRHOL', 'TRILC', 'TSKB', 'TSPOR', 'TTKOM', 'TTRAK', 'TUCLK', 'TUKAS', 'TUPRS', 'TUREX', 'TURGG', 'TURSG', 'UFUK', 'ULAS', 'ULKER', 'ULUFA', 'ULUSE', 'ULUUN', 'UNLU', 'USAK', 'USDTR', 'VAKBN', 'VAKFN', 'VAKKO', 'VANGD', 'VBTYZ', 'VERTU', 'VERUS', 'VESBE', 'VESTL', 'VKFYO', 'VKGYO', 'VKING', 'VRGYO', 'VSNMD', 'YAPRK', 'YATAS', 'YAYLA', 'YBTAS', 'YEOTK', 'YESIL', 'YGGYO', 'YGYO', 'YIGIT', 'YKBNK', 'YKSLN', 'YONGA', 'YUNSA', 'YYAPI', 'YYLGD', 'Z30EA', 'Z30KE', 'Z30KP', 'ZEDUR', 'ZELOT', 'ZGOLD', 'ZOREN', 'ZPBDL', 'ZPLIB', 'ZPT10', 'ZPX30', 'ZRE20', 'ZRGYO', 'ZSR25',
    # Eklediğin hisseler
    'MCARD', 'ZGYO', 'ZERGY', 'NETCD', 'ATATR'
]

valid_symbols = [s for s in symbols if len(s) >= 3 and s.isalpha() and s not in ['CUSAN', 'APMDL']]
bist_symbols = [s + '.IS' for s in valid_symbols]

print(f"{len(bist_symbols)} hisse taranıyor... (~8-15 dk)\n")

# ==================== PIVOT HIGH ====================
def ta_pivothigh(series, left, right):
    series = series.values.flatten()
    result = np.full(len(series), np.nan)
    for i in range(left, len(series) - right):
        if series[i] >= max(series[i-left:i]) and series[i] >= max(series[i+1:i+right+1]):
            result[i] = series[i]
    return pd.Series(result)

# ==================== 4 SAAT RESAMPLE ====================
def resample_to_4h(df):
    if df.empty: return df
    df = df.copy()
    df.index = pd.to_datetime(df.index)
    ohlc = {'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'}
    return df.resample('4H').agg(ohlc).dropna()

# ==================== İNDİKATÖRLER ====================
def calculate_indicators_daily(df):
    if tf["name"] == "4h":
        df = resample_to_4h(df)
    
    df = df.copy().reset_index(drop=True)
    high = df['High'].values.flatten()
    low = df['Low'].values.flatten()
    close = df['Close'].values.flatten()
    volume = df['Volume'].values.flatten()

    df['pivot_high'] = ta_pivothigh(df['High'], left_len, right_len)
    df['last_resistance'] = df['pivot_high'].ffill()
    df['donchian_high'] = pd.Series(high).rolling(donchian_period, min_periods=left_len).max()
    df['resistance_level'] = df['last_resistance'].fillna(df['donchian_high'])

    prev_close = np.roll(close, 1); prev_close[0] = close[0]
    tr = np.maximum.reduce([high - low, np.abs(high - prev_close), np.abs(low - prev_close)])
    df['atr'] = pd.Series(tr).rolling(atr_period, min_periods=1).mean()

    delta = np.diff(close, prepend=close[0])
    gain = np.where(delta > 0, delta, 0)
    loss = np.where(delta < 0, -delta, 0)
    gain_avg = pd.Series(gain).rolling(rsi_length, min_periods=1).mean()
    loss_avg = pd.Series(loss).rolling(rsi_length, min_periods=1).mean()
    rs = gain_avg / (loss_avg + 1e-10)
    df['rsi'] = 100 - (100 / (1 + rs))
    df['rsi_in_zone'] = (df['rsi'] < rsi_oversold) | (df['rsi'] > rsi_overbought)

    df['volume_avg'] = pd.Series(volume).rolling(volume_lookback, min_periods=1).mean()
    df['volume_increase'] = volume > df['volume_avg'] * 1.5

    momentum = pd.Series(close).diff(10)
    momentum_sma = momentum.rolling(10, min_periods=1).mean()
    df['momentum_strong'] = (momentum > 0) & (momentum > momentum_sma)

    plus_dm = np.diff(high, prepend=high[0]); plus_dm[plus_dm < 0] = 0
    minus_dm = np.diff(low, prepend=low[0]); minus_dm = -minus_dm; minus_dm[minus_dm < 0] = 0
    tr_smooth = pd.Series(tr).rolling(adx_length, min_periods=1).mean()
    plus_di = 100 * pd.Series(plus_dm).rolling(adx_length, min_periods=1).mean() / (tr_smooth + 1e-10)
    minus_di = 100 * pd.Series(minus_dm).rolling(adx_length, min_periods=1).mean() / (tr_smooth + 1e-10)
    dx = np.abs(plus_di - minus_di) / (plus_di + minus_di + 1e-10) * 100
    df['adx'] = dx.rolling(adx_length, min_periods=1).mean()
    df['adx_strong'] = df['adx'] > 25

    ema12 = pd.Series(close).ewm(span=12, adjust=False).mean()
    ema26 = pd.Series(close).ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    signal_line = macd_line.ewm(span=9, adjust=False).mean()
    df['macd_strong'] = (macd_line > signal_line) & (macd_line > 0)

    return df

# ==================== YAKINLIK ve YORUM (Değiştirilmedi) ====================
def get_yakinlik_kategori(gain_pct):
    if gain_pct <= 3:
        return f"{gain_pct:.1f}".replace('.', ','), "Çok Yakın"
    elif gain_pct <= 7:
        return f"{gain_pct:.1f}".replace('.', ','), "Yakın"
    elif gain_pct <= 15:
        return f"{gain_pct:.1f}".replace('.', ','), "Orta"
    else:
        return f"{gain_pct:.1f}".replace('.', ','), "Uzak"

def generate_yorum(guc, yakinlik_kategori, yakinlik_pct_str):
    pct = float(yakinlik_pct_str.replace(',', '.'))
    if guc >= 4.0:
        if pct <= 3.0:
            return "Direnci kırdı, artık direnç destek oldu. Kısa vadede kar realizasyonu gelebilir, stop seviyesi biraz altı takip edilmeli."
        elif pct <= 7.0:
            return "Güçlü kırılım gerçekleşti. Destek olarak çalışabilir. Kar satışları kısa vadede gelebilir, dikkatli olun."
        else:
            return "Çok güçlü kırılım ama fiyat bayağı uzaklaştı. Kar realizasyonunu unutmayın, kısa vadeli düzeltme gelebilir."
    elif guc >= 3.0:
        if pct <= 7.0:
            return "Kırılım onaylandı, direnç → destek dönüşümü olası. Kar realizasyonu riski orta seviyede. Stop seviyesi direncin biraz altına konulabilir."
        else:
            return "Kırılım gerçekleşti ancak mesafe açıldı. Kar satışları gelebilir, kısa vadeli dikkat gerekli."
    else:
        return "Kırılım eşiğinde güçlü sinyal. Direnç destek olabilir ama onay bekleniyor. Kar realizasyonu kısa vadede gelebilir."

# ==================== SİNYAL TESPİT (Değiştirilmedi) ====================
def detect_very_strong_breakout(df_daily, current_close, symbol):
    if len(df_daily) < 20:
        return None

    last_row = df_daily.iloc[-1]
    resistance = float(last_row['resistance_level'])
    if pd.isna(resistance) or resistance <= 0:
        return None

    if current_close < resistance * BREAKOUT_MULTIPLIER:
        return None

    rsi_ok = last_row['rsi_in_zone'].item()
    volume_ok = last_row['volume_increase'].item()
    atr_ok = float(last_row['atr']) > float(df_daily['atr'].iloc[-2]) if len(df_daily) > 1 else True
    momentum_ok = last_row['momentum_strong'].item()
    adx_ok = last_row['adx_strong'].item()
    macd_ok = last_row['macd_strong'].item()

    strength = 0.0
    if rsi_ok: strength += 1.5
    if volume_ok: strength += 1.5
    if atr_ok: strength += 1.0
    if momentum_ok: strength += 0.7
    if adx_ok: strength += 0.7
    if macd_ok: strength += 0.7

    count = sum([rsi_ok, volume_ok, atr_ok, momentum_ok, adx_ok, macd_ok])
    if count >= 3: strength = max(strength, 2.5)
    strength = min(strength, 5.0)

    if strength < MIN_STRENGTH:
        return None

    gain_pct = (current_close / resistance - 1) * 100
    uzaklik, kategori = get_yakinlik_kategori(gain_pct)

    key = symbol.replace('.IS', '')
    if key in broken_res and abs(broken_res[key] - resistance) < 1e-6:
        return None

    broken_res[key] = resistance

    yorum = generate_yorum(strength, kategori, uzaklik)

    return {
        'Sembol': key,
        'Tür': 'Kırılım AL',
        'Kategori': 'Çok Güçlü' if strength >= 4.0 else 'Güçlü',
        'Güç': round(strength, 1),
        'Fiyat': round(current_close, 2),
        'Direnç': round(resistance, 2),
        'Yakınlık %': uzaklik,
        'Yakınlık': kategori,
        'Yorum': yorum,
        'Periyot': tf["name"].upper()
    }

# ==================== TARAMA ====================
results = []
print(f"\nTarama başladı ({tf['name'].upper()}): {datetime.now().strftime('%d.%m.%Y %H:%M')}\n")

for i, symbol in enumerate(bist_symbols):
    try:
        data_daily = yf.download(symbol, period=tf["period"], interval=tf["interval"], progress=False, auto_adjust=True)
        if data_daily.empty or len(data_daily) < 20:
            continue

        current_close = float(data_daily['Close'].iloc[-1])
        data_daily = calculate_indicators_daily(data_daily)
        result = detect_very_strong_breakout(data_daily, current_close, symbol)
        if result:
            results.append(result)
            print(f"BULUNDU → {result['Sembol']} | Güç: {result['Güç']} | Yakınlık: {result['Yakınlık']} ({result['Yakınlık %']})")

        if (i + 1) % 50 == 0:
            print(f"[{i+1}/{len(bist_symbols)}] Bulunan: {len(results)}")

    except Exception as e:
        continue
    time.sleep(0.1)

# ==================== SONUÇLAR ====================
df_results = pd.DataFrame(results)

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

if not df_results.empty:
    df_results['Güç'] = "'" + df_results['Güç'].astype(str).str.replace('.', ',')
    df_results['Fiyat']   = "'" + df_results['Fiyat'].astype(float).round(2).astype(str).str.replace('.', ',')
    df_results['Direnç']  = "'" + df_results['Direnç'].astype(float).round(2).astype(str).str.replace('.', ',')

    df_results['Yakınlık_Num'] = df_results['Yakınlık %'].str.replace(',', '.').astype(float)
    df_results = df_results.sort_values(
        ['Güç', 'Yakınlık_Num'], ascending=[False, True]
    ).drop(columns=['Yakınlık_Num']).reset_index(drop=True)

    df_results['Yorum'] = df_results['Yorum'].str.wrap(60)

    top10 = df_results.head(10)
    esik = df_results[
        (df_results['Güç'].str.replace("'", '').str.replace(',', '.').astype(float) >= 2.0) &
        (df_results['Güç'].str.replace("'", '').str.replace(',', '.').astype(float) < 2.5)
    ].sort_values('Yakınlık %', ascending=True)

    print("\n" + "═"*140)
    print(f"İLK 10 ÇOK GÜÇLÜ / GÜÇLÜ KIRILIM ({tf['name'].upper()})".center(140))
    print("═"*140)
    if not top10.empty:
        print(top10[['Sembol','Güç','Fiyat','Direnç','Yakınlık %','Yakınlık','Yorum']].to_markdown(index=False))
    else:
        print("Henüz çok güçlü / güçlü kırılım yok.")

    print("\n" + "═"*140)
    print("KIRILIM EŞİĞİNDE OLANLAR (Güç 2.0-2.4)".center(140))
    print("═"*140)
    if not esik.empty:
        print(esik[['Sembol','Güç','Fiyat','Direnç','Yakınlık %','Yakınlık','Yorum']].to_markdown(index=False))
    else:
        print("Eşik seviyesinde aday yok.")

    # EXCEL
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename_xlsx = f"Pulse-TARAMA_{tf['name'].upper()}_{timestamp}.xlsx"
    full_path_xlsx = os.path.join(desktop_path, filename_xlsx)

    try:
        with pd.ExcelWriter(full_path_xlsx, engine='openpyxl') as writer:
            df_results[['Sembol','Tür','Kategori','Güç','Fiyat','Direnç','Yakınlık %','Yakınlık','Yorum','Periyot']].to_excel(
                writer, sheet_name='Kırılım Sinyalleri', index=False
            )
            worksheet = writer.sheets['Kırılım Sinyalleri']
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = (max_length + 2)

            worksheet.auto_filter.ref = worksheet.dimensions

        print(f"\nExcel kaydedildi → {full_path_xlsx}")
    except Exception as e:
        print(f"Excel hatası: {e}")

    # CSV Yedek
    filename_csv = f"Pulse_{tf['name'].upper()}_{timestamp}.csv"
    full_path_csv = os.path.join(desktop_path, filename_csv)
    df_results.to_csv(full_path_csv, index=False, encoding='utf-8-sig', sep=';')
    print(f"CSV kaydedildi → {full_path_csv}")

else:
    print("\nBu periyotta uygun sinyal bulunamadı.")

print("\nTarama tamamlandı.")
